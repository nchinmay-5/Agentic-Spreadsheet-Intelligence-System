"""Constrained LLM-assisted target-cell resolution."""

from __future__ import annotations

import re
from typing import Any, Protocol

from openpyxl.utils.cell import get_column_letter, range_boundaries


MAX_TABLE_CELLS = 400
MIN_CONFIDENCE = 0.85


class TableResolverAgent(Protocol):
    def rank_sheets(self, query: str, sheet_names: list[str]) -> list[dict[str, Any]]: ...
    def propose_table_ranges(self, query: str, sheet_context: dict[str, Any]) -> list[dict[str, Any]]: ...
    def confirm_table(self, query: str, table: dict[str, Any]) -> dict[str, Any]: ...
    def select_table_cell(self, query: str, table: dict[str, Any]) -> dict[str, Any]: ...


def build_search_index(workbook_index: dict[str, Any]) -> list[dict[str, Any]]:
    """Compatibility helper; table resolution intentionally has no global label index."""
    return []


def resolve_target(
    query: str,
    workbook_index: dict[str, Any],
    agent: TableResolverAgent,
) -> dict[str, Any]:
    """Resolve an explicit address or select a cell through bounded table evidence."""
    direct = _direct_reference(query.strip(), workbook_index)
    if direct:
        return {"status": "resolved", "node_id": direct, "method": "direct_reference", "candidates": []}

    sheets = {sheet["name"]: sheet for sheet in workbook_index["sheets"]}
    candidates: list[dict[str, Any]] = []
    for ranked in agent.rank_sheets(query, workbook_index["sheet_names"]):
        sheet = sheets.get(ranked.get("sheet"))
        if not sheet:
            continue
        for proposal in agent.propose_table_ranges(query, _sheet_context(sheet)):
            table = _extract_table(sheet, proposal.get("range"))
            if table is None:
                continue
            confirmation = agent.confirm_table(query, table)
            if not (confirmation.get("is_table") and confirmation.get("relevant_to_query") and _confidence(confirmation) >= MIN_CONFIDENCE):
                continue
            selection = agent.select_table_cell(query, table)
            address = selection.get("cell")
            if not isinstance(address, str) or address not in table["cell_addresses"] or _confidence(selection) < MIN_CONFIDENCE:
                continue
            return {
                "status": "resolved",
                "node_id": _node(sheet["name"], address),
                "method": "llm_confirmed_table",
                "candidates": candidates,
                "evidence": {"sheet": sheet["name"], "range": table["range"], "confirmation": confirmation, "selection": selection},
            }
        candidates.append({"sheet": sheet["name"], "sheet_confidence": ranked.get("confidence", 0)})
    return {"status": "unresolved", "node_id": None, "method": "llm_table_resolution", "candidates": candidates}


def _sheet_context(sheet: dict[str, Any]) -> dict[str, Any]:
    """Small structural summary for range proposal; do not send full numeric data yet."""
    text_cells = [
        {"cell": cell["address"], "text": cell["value"]}
        for cell in sheet["cells"]
        if isinstance(cell.get("value"), str) and cell["value"].strip()
    ][:200]
    numeric_or_formula_count = sum(
        cell.get("formula") is not None or isinstance(cell.get("value"), (int, float))
        for cell in sheet["cells"]
    )
    return {
        "sheet": sheet["name"],
        "used_range": sheet.get("used_range"),
        "text_cells": text_cells,
        "numeric_or_formula_cells": numeric_or_formula_count,
    }


def _extract_table(sheet: dict[str, Any], proposed_range: Any) -> dict[str, Any] | None:
    if not isinstance(proposed_range, str):
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(proposed_range.replace("$", ""))
    except ValueError:
        return None
    if min_row < 1 or min_col < 1 or max_row > sheet.get("max_row", 0) or max_col > sheet.get("max_column", 0):
        return None
    if (max_row - min_row + 1) * (max_col - min_col + 1) > MAX_TABLE_CELLS:
        return None
    cells = {cell["address"]: cell for cell in sheet["cells"]}
    rows: list[list[dict[str, Any]]] = []
    addresses: list[str] = []
    for row in range(min_row, max_row + 1):
        row_values = []
        for col in range(min_col, max_col + 1):
            address = f"{get_column_letter(col)}{row}"
            source = cells.get(address, {})
            row_values.append({"cell": address, "value": source.get("value"), "formula": source.get("formula")})
            addresses.append(address)
        rows.append(row_values)
    return {"sheet": sheet["name"], "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}", "rows": rows, "cell_addresses": addresses}


def _confidence(result: dict[str, Any]) -> float:
    try:
        value = float(result.get("confidence", 0))
    except (TypeError, ValueError):
        return 0.0
    return value if 0 <= value <= 1 else 0.0


def _direct_reference(text: str, workbook_index: dict[str, Any]) -> str | None:
    compact = text.replace("$", "")
    if "!" in compact:
        sheet, address = compact.rsplit("!", 1)
        sheet = sheet.strip("'")
        if re.fullmatch(r"[A-Za-z]{1,3}[1-9]\d*", address, re.I) and sheet in workbook_index["sheet_names"]:
            return _node(sheet, address.upper())
    for sheet in sorted(workbook_index["sheet_names"], key=len, reverse=True):
        prefix = f"{sheet} "
        if compact.lower().startswith(prefix.lower()):
            address = compact[len(prefix):].strip()
            if re.fullmatch(r"[A-Za-z]{1,3}[1-9]\d*", address, re.I):
                return _node(sheet, address.upper())
    return None


def _node(sheet: str, address: str) -> str:
    return f"'{sheet}'!{address}" if re.search(r"[^A-Za-z0-9_.]", sheet) else f"{sheet}!{address}"
