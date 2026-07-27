"""Create a JSON-safe index of workbook metadata and populated cells."""

from __future__ import annotations

import argparse
import json
from uuid import uuid4
from datetime import date, datetime, time
from numbers import Integral, Real
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from app.formulas.reference_parser import parse_references_with_diagnostics

def build_workbook_index(workbook_path: str | Path) -> dict[str, Any]:
    """Read an Excel workbook into a traceable, JSON-safe index.

    Formulas are loaded as their original Excel text. The function does not
    recalculate formula results.
    """

    source_path = Path(workbook_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {source_path}")

    if source_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError("Only modern Excel workbook formats are supported.")

    workbook = load_workbook(
        filename=source_path,
        read_only=False,
        data_only=False,
        keep_vba=source_path.suffix.lower() in {".xlsm", ".xltm"},
    )

    sheet_indexes = [_index_sheet(sheet) for sheet in workbook.worksheets]
    return {
        "workbook_id": _workbook_id(source_path),
        "source_file": source_path.name,
        "file_name": source_path.name,
        "sheet_names": workbook.sheetnames,
        "sheets": sheet_indexes,
    }


def index_workbook_to_file(
    workbook_path: str | Path, indexes_dir: str | Path = Path("data") / "indexes"
) -> tuple[dict[str, Any], Path]:
    """Build an index, save it by workbook ID, and return both values."""
    index = build_workbook_index(workbook_path)
    return index, save_workbook_index(index, indexes_dir)

def save_workbook_index(index: dict[str, Any], indexes_dir: str | Path) -> Path:
    """Persist an index as ``<workbook_id>.json`` and return its path."""
    workbook_id = _validated_workbook_id(index.get("workbook_id"))
    destination = Path(indexes_dir) / f"{workbook_id}.json"
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(index, indent=2, ensure_ascii=False), encoding="utf-8")
    return destination


def load_workbook_index(workbook_id: str, indexes_dir: str | Path) -> dict[str, Any]:
    """Load a persisted index by workbook ID."""
    safe_workbook_id = _validated_workbook_id(workbook_id)
    index_path = Path(indexes_dir) / f"{safe_workbook_id}.json"
    if not index_path.is_file():
        raise ValueError(f"Workbook not found: {safe_workbook_id}")

    try:
        index = json.loads(index_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise ValueError(f"Workbook index is invalid: {safe_workbook_id}") from error

    if not isinstance(index, dict) or index.get("workbook_id") != safe_workbook_id:
        raise ValueError(f"Workbook index does not match requested ID: {safe_workbook_id}")
    return index


def _validated_workbook_id(workbook_id: Any) -> str:
    if not isinstance(workbook_id, str) or not workbook_id:
        raise ValueError("Workbook ID is required.")
    if Path(workbook_id).name != workbook_id or workbook_id in {".", ".."}:
        raise ValueError("Invalid workbook ID.")
    return workbook_id


def _index_sheet(sheet: Any) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    formula_cells_count = 0
    value_cells_count = 0

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cells.append(_index_cell(cell))

    return {
        "name": sheet.title,
        "used_range": sheet.calculate_dimension(),
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_cells": [str(cell_range) for cell_range in sheet.merged_cells.ranges],
        "cells": cells,
    }


def _index_cell(cell: Cell) -> dict[str, Any]:
    is_formula = cell.data_type == "f"
    references, diagnostics = (
        parse_references_with_diagnostics(cell.value, cell.parent.title)
        if is_formula
        else ([], [])
    )
    return {
        "address": cell.coordinate,
        "value": None if is_formula else _json_value(cell.value),
        "formula": cell.value if is_formula else None,
        "references": references,
        # "reference_diagnostics": diagnostics,
        # "data_type": cell.data_type,
        "comment": cell.comment.text if cell.comment else None,
        "number_format": cell.number_format,
    }


def _workbook_id(source_path: Path) -> str:
    """Return a unique ID for this workbook upload."""
    return f"{source_path.stem}-{uuid4().hex}"


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, Integral):
        return int(value)
    if isinstance(value, Real):
        return float(value)
    if isinstance(value, (str, bool)) or value is None:
        return value
    return str(value)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a JSON index from an Excel workbook."
    )
    parser.add_argument("workbook", help="Path to the Excel workbook")
    parser.add_argument(
        "--output",
        help="JSON output path (default: data/indexes/<workbook name>.json)",
    )
    return parser.parse_args()

if __name__ == "__main__":
    args = _parse_args()
    written_path = index_workbook_to_file(args.workbook, args.output)
    work_ = Path(r"C:\Users\chinm\OneDrive\Desktop\Q_A_EXCEL\sample_data\Formula_Lineage_Test.xlsx")
    output_ = Path(r"C:\Users\chinm\OneDrive\Desktop\Q_A_EXCEL\sample_data\test1.txt")
    written_path = index_workbook_to_file(work_, output_)
    print(f"Workbook index written to {written_path}")
