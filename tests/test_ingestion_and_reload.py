import tempfile
from pathlib import Path
from unittest import TestCase

from openpyxl import Workbook

from app.graph.dependency_graph import build_dependency_graph, graph_from_json, graph_to_json
from app.ingestion.workbook_indexer import (
    build_workbook_index,
    index_workbook_to_file,
    load_workbook_index,
)


class IngestionAndReloadTests(TestCase):
    def test_index_and_graph_round_trip(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "book.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Input Sheet"
            sheet["A1"] = 10
            sheet["B1"] = "Revenue"
            sheet["C1"] = "=A1*2"
            workbook.save(path)
            index = build_workbook_index(path)
            graph = build_dependency_graph(index)
            restored = graph_from_json(graph_to_json(graph))
            self.assertTrue(index["workbook_id"].startswith("book-"))
            self.assertIn("'Input Sheet'!C1", restored)
            self.assertIn(("'Input Sheet'!A1", "'Input Sheet'!C1"), restored.edges)

    def test_does_not_serialize_fill_or_named_range_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "styled.xlsx"
            indexes_dir = Path(directory) / "indexes"
            workbook = Workbook()
            workbook.active["A1"] = "Styled value"
            workbook.save(path)
            index, output = index_workbook_to_file(path, indexes_dir)
            self.assertEqual(output, indexes_dir / f"{index['workbook_id']}.json")
            self.assertTrue(output.exists())
            self.assertNotIn("named_ranges", index)
            self.assertNotIn("fill", index["sheets"][0]["cells"][0])
            self.assertEqual(load_workbook_index(index["workbook_id"], indexes_dir), index)
